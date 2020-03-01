#
# genloadfiles.py - construct dutyman and google calendar upload files
#

import datetime as dt
from openpyxl import Workbook
from openpyxl import load_workbook

dmWorkbook = Workbook() # for dutyman
dmSheet = dmWorkbook.active
calWorkbook = Workbook() # for calendar
calSheet = calWorkbook.active

#
# function to check for ',' and then split and reverse the string
def checkname(name):
    sname = name
    #print (safety, type(safety))
    if (isinstance(name, str)):
        pos = name.find(",")
        if (pos > 0):
            sname = name[0:pos]
            cname = name[pos+2:]
            sname = cname + " " + sname
            #print (cname, sname)
        # strip any spaces at the end
        sname = sname.strip()

    return sname
# end checkname


# load the dutyman column lables
dmTitles = "Duty Date", "Duty Time", "Event", "Duty Type", "Swappable", \
        "Reminders", "Confirmed", "Duty Notify", "Duty Instructions", \
        "Duty DBID", "First Name", "Last Name", "Member Name", \
        "Name Qualifier", "Alloc", "Notes"

calTitles = "Start Date", "Start Time", "End Date", \
         "End Time", "Subject", "Description"

x = 1
for title in dmTitles:
        dmSheet.cell(row = 1, column = x).value = title
        x = x + 1
x = 1
for title in calTitles:
        calSheet.cell(row = 1, column = x).value = title
        x = x + 1

# open the scedule workbook
workbook = load_workbook(filename="RSC 2020 Schedule ver 270220.xlsx")
sheet = workbook.active

#print (workbook.sheetnames)
#print (sheet)
#print (sheet.title)

races = {} # blank list

x = 6 # offset to first data row

raceno = 0
for row in sheet.iter_rows(min_row=6, min_col=3, max_col=9, values_only=True):
    # is there a date for an 'event'
    if (isinstance(row[0], dt.date)):
	
        # put a check in for 'time' - if this is blank there is no race!
        if (isinstance(row[4], dt.time)):

            #safety = row[5]
            safety = checkname(row[5])
            #office = row[6]
            office = checkname(row[6])

            race = {
                "date": row[0],
                "hw": row[1],
                "lw": row[2],
                "name": row[3],
                "time": row[4],
                "safety": safety,
                "office": office
            }
            races[raceno] = race
            raceno = raceno + 1
        else:
            print ("Not scheduled event:",
                row[0].strftime("%d/%m/%Y"),
                row[3])

# list of types
dutytypes = [ {"duty": "safety", "name": "Safety Boat"},
    {"duty": "office", "name": "Race Office"}
]

# added to the working sheet
x = 2 # start adding at row after heading
for r in races:
    # guard as some entries are 'No Race" with a time
    if (races[r]["name"].find("No Race") == -1):
        #print (races[r]["date"], races[r]["name"], \
    	#    races[r]["time"], \
        #    races[r]["safety"],
        #    races[r]["office"])
        cname = "s" # defaults to flag issues
        sname = "s"

        for duty in dutytypes:
            dutyent = races[r][duty["duty"]] # is there a safety boat entry
            if (isinstance(dutyent, str) and len(dutyent) > 0):
                #print ("dutyent:", dutyent, len(dutyent))
                pos = dutyent.find(" ")
                if (pos >= 0):
                    cname = dutyent[0:pos]
                    sname = dutyent[pos+1:]
                else:
                    print ("No name", pos, dutyent)
                date = races[r]["date"].strftime("%d/%m/%Y")
                dmSheet.cell(row = x, column = 11).value = cname
                dmSheet.cell(row = x, column = 12).value = sname
                dmSheet.cell(row = x, column = 1).value = date
                dmSheet.cell(row = x, column = 2).value = races[r]["time"]
                dmSheet.cell(row = x, column = 3).value = races[r]["name"]
                dmSheet.cell(row = x, column = 4).value = duty["name"]
                x = x + 1
#
x = 2 # start adding at row after heading
for r in races:
    # load 'times'
    rt = races[r]["time"]
    lw = races[r]["lw"]
    hw = races[r]["hw"]

    if (races[r]["name"].find("No Race") == -1):
        desctide = ""
        if (isinstance(hw, dt.time)):
            tdiff = dt.timedelta(minutes=75)
            et = (dt.datetime.combine(dt.date.today(), rt) + tdiff).time()
            desctide = ". HW: " + hw.strftime("%H:%M")
        else:
            tdiff = dt.timedelta(minutes=120)
            et = (dt.datetime.combine(dt.date.today(), rt) + tdiff).time()
            if (isinstance(lw, dt.time)):
                desctide = ". LW: " + lw.strftime("%H:%M")

        #et = (dt.datetime.combine(dt.date.today(), rt) + tdiff).time()
        desc = ""
        for duty in dutytypes:
            dutyent = races[r][duty["duty"]] # is there a duty entry
            if (isinstance(dutyent, str) and len(dutyent) > 0):
                if (len(desc) > 0): desc = desc + ", "
                desc = desc + duty["name"] + ": " + dutyent

        if (len(desc) > 0): desc = desc + desctide
        #desc = desc + desctide
        date = races[r]["date"].strftime("%d/%m/%Y")
        stime = races[r]["time"].strftime("%H:%M")
        etime = et.strftime("%H:%M")
        calSheet.cell(row = x, column = 1).value = date
        calSheet.cell(row = x, column = 2).value = stime
        calSheet.cell(row = x, column = 3).value = date
        calSheet.cell(row = x, column = 4).value = etime
        calSheet.cell(row = x, column = 5).value = races[r]["name"]
        calSheet.cell(row = x, column = 6).value = desc

        x = x + 1

# and save the worksheets
dmWorkbook.save(filename="ranelagh-dutyman.xlsx")
calWorkbook.save(filename="ranelagh-google.xlsx")

# end of file
